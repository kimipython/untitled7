#!/usr/bin/python2
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from api_07_25.common.contents import case_url


class Read_Excel(object):

    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_excel(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        max_row = sheet.max_row
        max_cow = sheet.max_column
        data_list = []
        for row in range(2,max_row+1):
            for cow in range(1,max_cow+1):
                data_dict = {}
                data_dict[sheet.cell(1,cow).value] = sheet.cell(row,cow).value
                data_list.append(data_dict)
        wb.save(self.file_name)
        wb.close()
        return data_list

    def write_excel(self,row,actaul,result):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, 7).value = actaul
        sheet.cell(row, 8).value = result
        wb.save(self.file_name)
        wb.close()


if __name__ == '__main__':
    r = Read_Excel(case_url,'register')
    data = r.read_excel()
    # print(data)
    for a in data:
        print(a)
